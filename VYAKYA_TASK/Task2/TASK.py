import json
import datetime
import openpyxl
from openpyxl.workbook import Workbook
import pandas as pd

import xlsxwriter

wb = Workbook()
xlsxwriter.Workbook('out.xlsx')

ws1 = wb.create_sheet("Sheet_A")
ws1.title = "Deposit"

ws2 = wb.create_sheet("Sheet_B", 0)
ws2.title = "Withdrawals"

ws3 = wb.create_sheet("Sheet_C", 0)
ws3.title = "Insights"

wb.save(filename = 'out.xlsx')


with open("task_input_list.json") as blog_file:
    data = json.load(blog_file)
sheet_1 = 1
sheet_2 = 1
s=0

for i in range(77,133):
    #print(data[i])#3
    k=i+2
    try:
        datetime.datetime.strptime(data[i], '%m/%d/%y')


        m=1
        description=data[i+1]
        ws1.cell(row=1,column=6).value= 'Amount'
        ws1.cell(row=1,column=1).value= 'Date'
        ws1.cell(row=1,column=2).value= 'Day'
        ws1.cell(row=1,column=3).value= 'Month'
        ws1.cell(row=1,column=4).value= 'Year'
        ws1.cell(row=1,column=5).value= 'Description'

        ws2.cell(row=1,column=6).value= 'Amount'
        ws2.cell(row=1,column=1).value= 'Date'
        ws2.cell(row=1,column=2).value= 'Day'
        ws2.cell(row=1,column=3).value= 'Month'
        ws2.cell(row=1,column=4).value= 'Year'
        ws2.cell(row=1,column=5).value= 'Description'

        ws3.cell(row=1,column=1).value= 'Key'
        ws3.cell(row=1,column=2).value= 'Value'

        while m==1:
            adder=data[k]
            check=adder[0:2]
            if check.isalpha()==True:
                description=description+adder
                k=k+1

            else:
                m=0

                amount=data[k]
        ch1  = amount[0]
        if ch1 != "-" :
            print("greater :")
            sheet_1 = sheet_1 + 1
            dateSeperated = data[i].split("/")
            print(dateSeperated)
            ws1.cell(row=sheet_1,column=1).value= data[i]
            ws1.cell(row=sheet_1,column=2).value= dateSeperated[1]
            ws1.cell(row=sheet_1,column=3).value= dateSeperated[0]
            ws1.cell(row=sheet_1,column=4).value= dateSeperated[2]
            ws1.cell(row=sheet_1,column=5).value= description
            ws1.cell(row=sheet_1,column=6).value= amount
            wb.save(filename = 'out.xlsx')
            print("True")

        else :
            print("less")
            sheet_2 = sheet_2 + 1
            dateSeperated = data[i].split("/")
            print(dateSeperated)
            ws2.cell(row=sheet_2,column=1).value= data[i]
            ws2.cell(row=sheet_2,column=2).value= dateSeperated[1]
            ws2.cell(row=sheet_2,column=3).value= dateSeperated[0]
            ws2.cell(row=sheet_2,column=4).value= dateSeperated[2]
            ws2.cell(row=sheet_2,column=5).value= description
            ws2.cell(row=sheet_2,column=6).value= amount
            wb.save(filename = 'out.xlsx')
            print("True")
        
        user=ws1.cell(row=2,column=6).value

        print(user)
        
       



    except:
        #print("Incorrect data format, should be YYYY-MM-DD")
        pass

xls = pd.ExcelFile('out.xlsx')
df2 = pd.read_excel(xls, 'Withdrawals')
df1 = pd.read_excel(xls, 'Deposit')

maxa = max(df1['Amount'])
df2.sort_values(by='Amount', ascending=False)
leno = len(df2)
mini =  df2['Amount'][leno-1]
print(df2['Amount'][leno-1])
print(maxa)
ws3.cell(row=5,column=1).value= 'Maximum_amount'
ws3.cell(row=5,column=2).value= maxa
wb.save(filename = 'out.xlsx')

ws3.cell(row=6,column=1).value= 'Minimum_amount'
ws3.cell(row=6,column=2).value= mini
wb.save(filename = 'out.xlsx')
flag = 0
for i in range(2, 77) :
    web = str(data[i])
    if web.find(".com") != -1 and flag != 1:
        print(data[i])
        ws3.cell(row=2,column=1).value= 'Website'
        ws3.cell(row=2,column=2).value= data[i]
        wb.save(filename = 'out.xlsx')
        flag = 1

mail = []
mailone = ''

for i in range(2, 77) :
    web = str(data[i])
    if web.find("@gmail.com") != -1 :
        mailone = mailone + data[i]
        
        mail.append(data[i])
        ws3.cell(row=3,column=1).value= 'E-mail'
        ws3.cell(row=3,column=2).value= mailone
        wb.save(filename = 'out.xlsx')

if len(mail) == 0:
    ws3.cell(row=3,column=1).value= 'E-mail'
    ws3.cell(row=3,column=2).value= 'NA'
    wb.save(filename = 'out.xlsx')


   
num = ''
for i in range(6, 9) :
    n1 = data[i].split(':')
    num =  num + n1[1] + ',' 
    print(num)
ws3.cell(row=4,column=1).value= 'phone_number'
ws3.cell(row=4,column=2).value= num
wb.save(filename = 'out.xlsx')


